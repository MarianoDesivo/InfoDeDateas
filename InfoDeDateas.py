import requests
import bs4
import openpyxl
import re

wb = openpyxl.load_workbook(r'c:\users\Mariano\desktop\Nombres.xlsx') #Path de un excel con lista de nombres
hoja = wb['Sistemas'] #nombre de la pestaña del excel con la lista de nombres
nombre=[]
hoja2 = wb.create_sheet() #crear una nueva pestaña con los cuiles obtenidos, ademas del nombre al que pertenecen
hoja2.title='cuiles y nombres' #Cambio de nombre a la pestaña creada
cuiles = wb['cuiles y nombres'] #Se crea un objeto con la pestaña creada anteriormente
indiceexcel=1

for i in range(hoja.max_row):            
        nombre.append(str.split(str((hoja.cell(i,2).value)))) #los nombres que estan en la columna 2 van a la lista "nombre"

        
for i in range(len(nombre)):
    try: 
                nombreurl='+'.join(nombre[i]) #Para el url de dateas hay que poner un + entre nombres y apellido
                url = 'https://www.dateas.com/es/consulta_cuit_cuil?name='+ nombreurl + '&cuit=&tipo=personas-fisicas' #Preramos el url
                headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'} #Declaramos el User Agent del "navegador"
                result = requests.get(url, headers=headers) #Obtenemos la informacion de la pagina
                
                #Imprimimos algunas variables útiles para debug
                print(result.content.decode())
                print('i='+ str(i))
                print(result.raise_for_status)
                print(result.status_code)

                for j in range(20): #Cada pagina de dateas tiene 20 resultados
                    if nombre[i] == '': #Si la casilla del excel estaba vacía se saltea
                        continue
                    soup = bs4.BeautifulSoup(result.text, 'html.parser') #Creamos un objeto de Beautiful Soup
                    elems = soup.select('#mainContent > table > tbody > tr:nth-child(' + str(j+1) + ') > td:nth-child(1) > a') #Se guarda una lista con todos los nombres que muestra Dateas


                    if sorted(elems[0].text.split()) == sorted(nombre[i]): #Si el nombre que muestra dateas es igual al nombre del excel:
                        print(str(nombre[i]) + ' este lo encontre en ' + str(j+1)) #Imprime el nombre y en qué fila de dateas estaba
                        elemcuil = soup.select('#mainContent > table > tbody > tr:nth-child(' + str(j+1) + ') > td:nth-child(2) > a') #Se guarda su Cuil en una variable
                        print(elemcuil[0].text)
                        cuil= elemcuil[0].text.replace('-','') #Dateas usa guiones en los cuiles, los borramos
                        print(cuil)

                        #Escribimos en nuestro archivo excel el numero de cuil junto con el nombre de la persona
                        cuiles.cell(indiceexcel,1).value=cuil
                        cuiles.cell(indiceexcel,2).value=' '.join(nombre[i])
                        indiceexcel += 1
                        print(str(indiceexcel) + '--------------------------')
                        
    except:
            pass
wb.save('CUILES.xlsx') #Guardamos el excel completo
